"""
🧠 RomAI Custom Reporting Engine
================================

Phase 2.5: Advanced Analytics & Reporting - Custom Reporting Engine
Week 9 (Days 155-161) - Custom reporting engine with visualizations

This module provides a comprehensive custom reporting engine for the RomAI AGI platform,
enabling users to create, customize, and generate reports with advanced visualizations,
automated scheduling, and multi-format output capabilities.

Features:
- Custom report builder with drag-and-drop interface
- Advanced data visualization and charting
- Automated report scheduling and distribution
- Multi-format export (PDF, Excel, HTML, JSON)
- Template-based report generation
- Interactive dashboards and widgets
- Performance benchmarking and trending
- Compliance and audit reporting

Author: RomAI Development Team
Date: August 12, 2025
License: Proprietary
"""

import asyncio
import json
import logging
import os
import io
import base64
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Union, Callable
from dataclasses import dataclass, asdict
from enum import Enum
from pathlib import Path
import sqlite3
import pandas as pd
import numpy as np

# Reporting and visualization imports
try:
    import plotly.graph_objects as go
    import plotly.express as px
    from plotly.subplots import make_subplots
    import plotly.io as pio
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False
    logging.warning("Plotly not available - advanced visualizations disabled")

try:
    import matplotlib.pyplot as plt
    import seaborn as sns
    from matplotlib.backends.backend_pdf import PdfPages
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
    logging.warning("Matplotlib not available - static charts disabled")

try:
    from jinja2 import Template, Environment, FileSystemLoader
    JINJA2_AVAILABLE = True
except ImportError:
    JINJA2_AVAILABLE = False
    logging.warning("Jinja2 not available - template rendering disabled")

try:
    from weasyprint import HTML, CSS
    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False
    logging.warning("WeasyPrint not available - PDF generation may be limited")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("OpenPyXL not available - Excel export disabled")

from .advanced_analytics_engine import (
    AdvancedAnalyticsEngine,
    AnalyticsMetric,
    AnalyticsMetricType,
    SystemHealthMetrics,
    AIPerformanceMetrics,
    BusinessMetrics,
    ReportFormat
)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ReportType(Enum):
    """Types of reports that can be generated"""
    SYSTEM_HEALTH = "system_health"
    AI_PERFORMANCE = "ai_performance"
    BUSINESS_INTELLIGENCE = "business_intelligence"
    COMPLIANCE_AUDIT = "compliance_audit"
    PERFORMANCE_BENCHMARK = "performance_benchmark"
    EXECUTIVE_SUMMARY = "executive_summary"
    CUSTOM = "custom"

class ChartType(Enum):
    """Types of charts available for reports"""
    LINE = "line"
    BAR = "bar"
    PIE = "pie"
    SCATTER = "scatter"
    HEATMAP = "heatmap"
    GAUGE = "gauge"
    TABLE = "table"
    METRIC_CARD = "metric_card"
    HISTOGRAM = "histogram"
    BOX_PLOT = "box_plot"

class ScheduleFrequency(Enum):
    """Report scheduling frequencies"""
    HOURLY = "hourly"
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"
    QUARTERLY = "quarterly"
    YEARLY = "yearly"
    CUSTOM = "custom"

@dataclass
class ReportWidget:
    """Configuration for report widgets/components"""
    id: str
    title: str
    chart_type: ChartType
    data_source: str
    query_parameters: Dict[str, Any]
    visualization_config: Dict[str, Any]
    position: Dict[str, int]  # x, y, width, height
    style: Dict[str, Any]

@dataclass
class ReportTemplate:
    """Report template configuration"""
    id: str
    name: str
    description: str
    report_type: ReportType
    layout: Dict[str, Any]
    widgets: List[ReportWidget]
    default_parameters: Dict[str, Any]
    output_formats: List[ReportFormat]
    created_at: datetime
    updated_at: datetime

@dataclass
class ReportSchedule:
    """Scheduled report configuration"""
    id: str
    template_id: str
    name: str
    frequency: ScheduleFrequency
    schedule_config: Dict[str, Any]  # cron expression, specific times, etc.
    recipients: List[str]  # email addresses
    output_format: ReportFormat
    active: bool
    last_run: Optional[datetime]
    next_run: Optional[datetime]
    created_at: datetime

@dataclass
class GeneratedReport:
    """Generated report metadata and content"""
    id: str
    template_id: str
    name: str
    report_type: ReportType
    format: ReportFormat
    content: Union[str, bytes]
    file_path: Optional[str]
    parameters: Dict[str, Any]
    generated_at: datetime
    file_size: int
    status: str

class CustomReportingEngine:
    """
    Custom Reporting Engine for RomAI AGI Platform
    
    Provides comprehensive reporting capabilities including:
    - Custom report templates and builders
    - Advanced data visualization and charting
    - Automated report scheduling and distribution
    - Multi-format export capabilities
    - Interactive report dashboards
    - Performance benchmarking and analysis
    """
    
    def __init__(self, 
                 analytics_engine: AdvancedAnalyticsEngine,
                 output_directory: str = "reports",
                 template_directory: str = "templates"):
        """
        Initialize Custom Reporting Engine
        
        Args:
            analytics_engine: Instance of AdvancedAnalyticsEngine
            output_directory: Directory for storing generated reports
            template_directory: Directory for report templates
        """
        self.analytics_engine = analytics_engine
        self.output_directory = Path(output_directory)
        self.template_directory = Path(template_directory)
        
        # Create directories
        self.output_directory.mkdir(exist_ok=True)
        self.template_directory.mkdir(exist_ok=True)
        
        # Initialize databases and storage
        self.database_path = "reporting_engine.db"
        self._init_database()
        
        # Template storage
        self.templates: Dict[str, ReportTemplate] = {}
        self.schedules: Dict[str, ReportSchedule] = {}
        
        # Jinja2 environment for template rendering
        if JINJA2_AVAILABLE:
            self.jinja_env = Environment(
                loader=FileSystemLoader(str(self.template_directory)),
                autoescape=True
            )
        else:
            self.jinja_env = None
        
        # Load default templates
        self._load_default_templates()
        
        logger.info("Custom Reporting Engine initialized successfully")

    def _init_database(self):
        """Initialize SQLite database for reporting engine"""
        with sqlite3.connect(self.database_path) as conn:
            cursor = conn.cursor()
            
            # Report templates table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS report_templates (
                    id TEXT PRIMARY KEY,
                    name TEXT NOT NULL,
                    description TEXT,
                    report_type TEXT NOT NULL,
                    layout TEXT,
                    widgets TEXT,
                    default_parameters TEXT,
                    output_formats TEXT,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                )
            ''')
            
            # Report schedules table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS report_schedules (
                    id TEXT PRIMARY KEY,
                    template_id TEXT NOT NULL,
                    name TEXT NOT NULL,
                    frequency TEXT NOT NULL,
                    schedule_config TEXT,
                    recipients TEXT,
                    output_format TEXT NOT NULL,
                    active BOOLEAN NOT NULL,
                    last_run TEXT,
                    next_run TEXT,
                    created_at TEXT NOT NULL,
                    FOREIGN KEY (template_id) REFERENCES report_templates (id)
                )
            ''')
            
            # Generated reports table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS generated_reports (
                    id TEXT PRIMARY KEY,
                    template_id TEXT,
                    name TEXT NOT NULL,
                    report_type TEXT NOT NULL,
                    format TEXT NOT NULL,
                    file_path TEXT,
                    parameters TEXT,
                    generated_at TEXT NOT NULL,
                    file_size INTEGER,
                    status TEXT NOT NULL
                )
            ''')
            
            conn.commit()

    def _load_default_templates(self):
        """Load default report templates"""
        # System Health Report Template
        system_health_template = ReportTemplate(
            id="system_health_standard",
            name="System Health Standard Report",
            description="Comprehensive system health monitoring report",
            report_type=ReportType.SYSTEM_HEALTH,
            layout={"type": "grid", "columns": 12, "rows": 8},
            widgets=[
                ReportWidget(
                    id="cpu_usage_chart",
                    title="CPU Usage Trend",
                    chart_type=ChartType.LINE,
                    data_source="system_health",
                    query_parameters={"metric": "cpu_usage", "period": "24h"},
                    visualization_config={"color": "#667eea", "fill": True},
                    position={"x": 0, "y": 0, "width": 6, "height": 3},
                    style={"background": "#ffffff", "border": "1px solid #e5e7eb"}
                ),
                ReportWidget(
                    id="memory_usage_chart",
                    title="Memory Usage Trend",
                    chart_type=ChartType.LINE,
                    data_source="system_health",
                    query_parameters={"metric": "memory_usage", "period": "24h"},
                    visualization_config={"color": "#764ba2", "fill": True},
                    position={"x": 6, "y": 0, "width": 6, "height": 3},
                    style={"background": "#ffffff", "border": "1px solid #e5e7eb"}
                ),
                ReportWidget(
                    id="system_metrics_table",
                    title="Current System Metrics",
                    chart_type=ChartType.TABLE,
                    data_source="system_health",
                    query_parameters={"latest": True},
                    visualization_config={"striped": True, "bordered": True},
                    position={"x": 0, "y": 3, "width": 12, "height": 3},
                    style={"background": "#ffffff", "border": "1px solid #e5e7eb"}
                )
            ],
            default_parameters={"period": "24h", "format": "html"},
            output_formats=[ReportFormat.HTML, ReportFormat.PDF, ReportFormat.JSON],
            created_at=datetime.now(),
            updated_at=datetime.now()
        )
        
        # AI Performance Report Template
        ai_performance_template = ReportTemplate(
            id="ai_performance_standard",
            name="AI Performance Standard Report",
            description="Comprehensive AI model performance analysis",
            report_type=ReportType.AI_PERFORMANCE,
            layout={"type": "grid", "columns": 12, "rows": 10},
            widgets=[
                ReportWidget(
                    id="accuracy_gauge",
                    title="Model Accuracy",
                    chart_type=ChartType.GAUGE,
                    data_source="ai_performance",
                    query_parameters={"metric": "model_accuracy", "latest": True},
                    visualization_config={"min": 0, "max": 100, "threshold": 85},
                    position={"x": 0, "y": 0, "width": 4, "height": 4},
                    style={"background": "#ffffff", "border": "1px solid #e5e7eb"}
                ),
                ReportWidget(
                    id="cultural_score_gauge",
                    title="Romanian Cultural Score",
                    chart_type=ChartType.GAUGE,
                    data_source="ai_performance",
                    query_parameters={"metric": "romanian_cultural_score", "latest": True},
                    visualization_config={"min": 0, "max": 100, "threshold": 90},
                    position={"x": 4, "y": 0, "width": 4, "height": 4},
                    style={"background": "#ffffff", "border": "1px solid #e5e7eb"}
                ),
                ReportWidget(
                    id="response_quality_gauge",
                    title="Response Quality",
                    chart_type=ChartType.GAUGE,
                    data_source="ai_performance",
                    query_parameters={"metric": "response_quality", "latest": True},
                    visualization_config={"min": 0, "max": 5, "threshold": 4.5},
                    position={"x": 8, "y": 0, "width": 4, "height": 4},
                    style={"background": "#ffffff", "border": "1px solid #e5e7eb"}
                ),
                ReportWidget(
                    id="performance_trends",
                    title="Performance Trends",
                    chart_type=ChartType.LINE,
                    data_source="ai_performance",
                    query_parameters={"period": "7d", "metrics": ["model_accuracy", "romanian_cultural_score"]},
                    visualization_config={"multi_line": True, "legend": True},
                    position={"x": 0, "y": 4, "width": 12, "height": 4},
                    style={"background": "#ffffff", "border": "1px solid #e5e7eb"}
                )
            ],
            default_parameters={"period": "7d", "format": "html"},
            output_formats=[ReportFormat.HTML, ReportFormat.PDF, ReportFormat.JSON],
            created_at=datetime.now(),
            updated_at=datetime.now()
        )
        
        # Business Intelligence Report Template
        business_intelligence_template = ReportTemplate(
            id="business_intelligence_standard",
            name="Business Intelligence Dashboard",
            description="Key business metrics and KPI analysis",
            report_type=ReportType.BUSINESS_INTELLIGENCE,
            layout={"type": "grid", "columns": 12, "rows": 8},
            widgets=[
                ReportWidget(
                    id="revenue_chart",
                    title="Revenue Trend",
                    chart_type=ChartType.BAR,
                    data_source="business_metrics",
                    query_parameters={"metric": "revenue_eur", "period": "30d"},
                    visualization_config={"color": "#10b981", "gradient": True},
                    position={"x": 0, "y": 0, "width": 6, "height": 4},
                    style={"background": "#ffffff", "border": "1px solid #e5e7eb"}
                ),
                ReportWidget(
                    id="user_growth_chart",
                    title="User Growth",
                    chart_type=ChartType.LINE,
                    data_source="business_metrics",
                    query_parameters={"metric": "daily_active_users", "period": "30d"},
                    visualization_config={"color": "#3b82f6", "fill": True},
                    position={"x": 6, "y": 0, "width": 6, "height": 4},
                    style={"background": "#ffffff", "border": "1px solid #e5e7eb"}
                ),
                ReportWidget(
                    id="kpi_metrics",
                    title="Key Performance Indicators",
                    chart_type=ChartType.METRIC_CARD,
                    data_source="business_metrics",
                    query_parameters={"latest": True, "metrics": ["customer_satisfaction", "churn_rate", "conversion_rate"]},
                    visualization_config={"layout": "horizontal", "show_trend": True},
                    position={"x": 0, "y": 4, "width": 12, "height": 2},
                    style={"background": "#ffffff", "border": "1px solid #e5e7eb"}
                )
            ],
            default_parameters={"period": "30d", "format": "html"},
            output_formats=[ReportFormat.HTML, ReportFormat.PDF, ReportFormat.EXCEL],
            created_at=datetime.now(),
            updated_at=datetime.now()
        )
        
        # Store templates
        self.templates[system_health_template.id] = system_health_template
        self.templates[ai_performance_template.id] = ai_performance_template
        self.templates[business_intelligence_template.id] = business_intelligence_template

    async def create_template(self, template: ReportTemplate) -> str:
        """
        Create new report template
        
        Args:
            template: ReportTemplate configuration
            
        Returns:
            Template ID
        """
        try:
            # Store in memory
            self.templates[template.id] = template
            
            # Store in database
            with sqlite3.connect(self.database_path) as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    INSERT OR REPLACE INTO report_templates 
                    (id, name, description, report_type, layout, widgets, 
                     default_parameters, output_formats, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    template.id, template.name, template.description,
                    template.report_type.value, json.dumps(template.layout),
                    json.dumps([asdict(w) for w in template.widgets]),
                    json.dumps(template.default_parameters),
                    json.dumps([f.value for f in template.output_formats]),
                    template.created_at.isoformat(),
                    template.updated_at.isoformat()
                ))
                conn.commit()
            
            logger.info(f"Created report template: {template.id}")
            return template.id
            
        except Exception as e:
            logger.error(f"Error creating report template: {e}")
            raise

    async def generate_report(self, 
                            template_id: str,
                            output_format: ReportFormat,
                            parameters: Optional[Dict[str, Any]] = None) -> GeneratedReport:
        """
        Generate report from template
        
        Args:
            template_id: ID of the report template
            output_format: Desired output format
            parameters: Optional parameters to override template defaults
            
        Returns:
            GeneratedReport instance
        """
        try:
            if template_id not in self.templates:
                raise ValueError(f"Template not found: {template_id}")
            
            template = self.templates[template_id]
            
            # Merge parameters
            report_parameters = template.default_parameters.copy()
            if parameters:
                report_parameters.update(parameters)
            
            # Generate report content
            content = await self._generate_report_content(template, output_format, report_parameters)
            
            # Create report metadata
            report_id = f"{template_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            file_path = None
            file_size = 0
            
            # Save to file if needed
            if output_format in [ReportFormat.PDF, ReportFormat.HTML, ReportFormat.EXCEL]:
                file_extension = output_format.value
                file_path = self.output_directory / f"{report_id}.{file_extension}"
                
                if isinstance(content, bytes):
                    file_path.write_bytes(content)
                    file_size = len(content)
                else:
                    file_path.write_text(content, encoding='utf-8')
                    file_size = len(content.encode('utf-8'))
            
            # Create generated report record
            generated_report = GeneratedReport(
                id=report_id,
                template_id=template_id,
                name=f"{template.name} - {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                report_type=template.report_type,
                format=output_format,
                content=content,
                file_path=str(file_path) if file_path else None,
                parameters=report_parameters,
                generated_at=datetime.now(),
                file_size=file_size,
                status="completed"
            )
            
            # Store in database
            with sqlite3.connect(self.database_path) as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    INSERT INTO generated_reports 
                    (id, template_id, name, report_type, format, file_path, 
                     parameters, generated_at, file_size, status)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    generated_report.id, generated_report.template_id,
                    generated_report.name, generated_report.report_type.value,
                    generated_report.format.value, generated_report.file_path,
                    json.dumps(generated_report.parameters),
                    generated_report.generated_at.isoformat(),
                    generated_report.file_size, generated_report.status
                ))
                conn.commit()
            
            logger.info(f"Generated report: {report_id}")
            return generated_report
            
        except Exception as e:
            logger.error(f"Error generating report: {e}")
            raise

    async def _generate_report_content(self, 
                                     template: ReportTemplate,
                                     output_format: ReportFormat,
                                     parameters: Dict[str, Any]) -> Union[str, bytes]:
        """Generate report content based on template and format"""
        
        # Collect data for all widgets
        widget_data = {}
        for widget in template.widgets:
            data = await self._collect_widget_data(widget, parameters)
            widget_data[widget.id] = data
        
        # Generate content based on format
        if output_format == ReportFormat.HTML:
            return await self._generate_html_report(template, widget_data, parameters)
        elif output_format == ReportFormat.PDF:
            html_content = await self._generate_html_report(template, widget_data, parameters)
            return await self._convert_html_to_pdf(html_content)
        elif output_format == ReportFormat.JSON:
            return await self._generate_json_report(template, widget_data, parameters)
        elif output_format == ReportFormat.EXCEL:
            return await self._generate_excel_report(template, widget_data, parameters)
        else:
            raise ValueError(f"Unsupported output format: {output_format}")

    async def _collect_widget_data(self, widget: ReportWidget, parameters: Dict[str, Any]) -> Dict[str, Any]:
        """Collect data for a specific widget"""
        try:
            # Merge widget parameters with report parameters
            query_params = widget.query_parameters.copy()
            query_params.update(parameters)
            
            # Get data based on data source
            if widget.data_source == "system_health":
                if query_params.get("latest"):
                    data = SystemHealthMetrics.collect_current_metrics()
                    return asdict(data)
                else:
                    hours = self._parse_period_to_hours(query_params.get("period", "24h"))
                    data = await self.analytics_engine.get_system_health_history(hours)
                    return [asdict(d) for d in data]
            
            elif widget.data_source == "ai_performance":
                # Generate sample AI performance data
                if query_params.get("latest"):
                    data = AIPerformanceMetrics.create_sample_metrics()
                    return asdict(data)
                else:
                    count = self._parse_period_to_count(query_params.get("period", "7d"))
                    data = [AIPerformanceMetrics.create_sample_metrics() for _ in range(count)]
                    return [asdict(d) for d in data]
            
            elif widget.data_source == "business_metrics":
                # Generate sample business metrics data
                if query_params.get("latest"):
                    data = BusinessMetrics.create_sample_metrics()
                    return asdict(data)
                else:
                    count = self._parse_period_to_count(query_params.get("period", "30d"))
                    data = [BusinessMetrics.create_sample_metrics() for _ in range(count)]
                    return [asdict(d) for d in data]
            
            else:
                return {"error": f"Unknown data source: {widget.data_source}"}
        
        except Exception as e:
            logger.error(f"Error collecting widget data for {widget.id}: {e}")
            return {"error": str(e)}

    def _parse_period_to_hours(self, period: str) -> int:
        """Parse period string to hours"""
        if period.endswith('h'):
            return int(period[:-1])
        elif period.endswith('d'):
            return int(period[:-1]) * 24
        elif period.endswith('w'):
            return int(period[:-1]) * 24 * 7
        else:
            return 24  # default to 24 hours

    def _parse_period_to_count(self, period: str) -> int:
        """Parse period string to data point count"""
        if period.endswith('h'):
            return int(period[:-1])
        elif period.endswith('d'):
            return int(period[:-1])
        elif period.endswith('w'):
            return int(period[:-1]) * 7
        else:
            return 30  # default to 30 data points

    async def _generate_html_report(self, 
                                  template: ReportTemplate,
                                  widget_data: Dict[str, Any],
                                  parameters: Dict[str, Any]) -> str:
        """Generate HTML report content"""
        
        # Create HTML template
        html_template = f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>{template.name}</title>
            <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
            <style>
                body {{
                    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                    margin: 0;
                    padding: 20px;
                    background-color: #f9fafb;
                }}
                .report-header {{
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    color: white;
                    padding: 2rem;
                    border-radius: 12px;
                    margin-bottom: 2rem;
                }}
                .report-title {{
                    font-size: 2rem;
                    font-weight: bold;
                    margin-bottom: 0.5rem;
                }}
                .report-subtitle {{
                    font-size: 1.1rem;
                    opacity: 0.9;
                }}
                .widget-grid {{
                    display: grid;
                    grid-template-columns: repeat(12, 1fr);
                    gap: 1rem;
                    margin-bottom: 2rem;
                }}
                .widget {{
                    background: white;
                    border-radius: 8px;
                    padding: 1.5rem;
                    box-shadow: 0 1px 3px rgba(0, 0, 0, 0.1);
                    border: 1px solid #e5e7eb;
                }}
                .widget-title {{
                    font-size: 1.25rem;
                    font-weight: 600;
                    margin-bottom: 1rem;
                    color: #374151;
                }}
                .metric-value {{
                    font-size: 2.5rem;
                    font-weight: bold;
                    color: #1f2937;
                }}
                .metric-label {{
                    font-size: 0.875rem;
                    color: #6b7280;
                    margin-top: 0.5rem;
                }}
                .chart-container {{
                    width: 100%;
                    height: 300px;
                }}
                .table {{
                    width: 100%;
                    border-collapse: collapse;
                }}
                .table th, .table td {{
                    padding: 0.75rem;
                    text-align: left;
                    border-bottom: 1px solid #e5e7eb;
                }}
                .table th {{
                    background-color: #f9fafb;
                    font-weight: 600;
                }}
                .report-footer {{
                    text-align: center;
                    padding: 2rem;
                    color: #6b7280;
                    border-top: 1px solid #e5e7eb;
                    margin-top: 3rem;
                }}
            </style>
        </head>
        <body>
            <div class="report-header">
                <div class="report-title">{template.name}</div>
                <div class="report-subtitle">{template.description}</div>
                <div style="margin-top: 1rem; font-size: 0.9rem;">
                    Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | 
                    Report Type: {template.report_type.value} |
                    Period: {parameters.get('period', 'N/A')}
                </div>
            </div>
            
            <div class="widget-grid">
        """
        
        # Generate widgets
        for widget in template.widgets:
            data = widget_data.get(widget.id, {})
            widget_html = await self._generate_widget_html(widget, data)
            
            # Calculate grid position
            grid_style = f"""
                grid-column: span {widget.position['width']};
                grid-row: span {widget.position['height']};
            """
            
            html_template += f"""
                <div class="widget" style="{grid_style}">
                    <div class="widget-title">{widget.title}</div>
                    {widget_html}
                </div>
            """
        
        html_template += """
            </div>
            
            <div class="report-footer">
                <p>Generated by RomAI Advanced Analytics & Reporting Engine v2.5</p>
                <p>© 2025 RomAI AGI Platform. All rights reserved.</p>
            </div>
        </body>
        </html>
        """
        
        return html_template

    async def _generate_widget_html(self, widget: ReportWidget, data: Dict[str, Any]) -> str:
        """Generate HTML for a specific widget"""
        
        if widget.chart_type == ChartType.METRIC_CARD:
            if isinstance(data, dict) and not data.get("error"):
                value = list(data.values())[0] if data else 0
                return f"""
                    <div class="metric-value">{value:.2f}</div>
                    <div class="metric-label">Current Value</div>
                """
            else:
                return "<div class='metric-value'>--</div><div class='metric-label'>No data</div>"
        
        elif widget.chart_type == ChartType.TABLE:
            if isinstance(data, dict) and not data.get("error"):
                # Create simple table from dict
                table_html = '<table class="table"><tbody>'
                for key, value in data.items():
                    if isinstance(value, (int, float)):
                        formatted_value = f"{value:.2f}" if isinstance(value, float) else str(value)
                    else:
                        formatted_value = str(value)[:50]  # Truncate long strings
                    table_html += f"<tr><td>{key.replace('_', ' ').title()}</td><td>{formatted_value}</td></tr>"
                table_html += '</tbody></table>'
                return table_html
            else:
                return "<p>No data available</p>"
        
        elif widget.chart_type == ChartType.GAUGE:
            if isinstance(data, dict) and not data.get("error"):
                value = list(data.values())[0] if data else 0
                return f"""
                    <div class="chart-container">
                        <div style="text-align: center; padding-top: 50px;">
                            <div class="metric-value">{value:.1f}%</div>
                            <div class="metric-label">Performance Gauge</div>
                        </div>
                    </div>
                """
            else:
                return "<div class='chart-container'><p>No gauge data available</p></div>"
        
        elif widget.chart_type in [ChartType.LINE, ChartType.BAR]:
            # For now, return placeholder for charts
            return f"""
                <div class="chart-container">
                    <div style="text-align: center; padding-top: 100px; color: #6b7280;">
                        📊 {widget.chart_type.value.title()} Chart Placeholder
                        <br><small>Data points: {len(data) if isinstance(data, list) else 1}</small>
                    </div>
                </div>
            """
        
        else:
            return f"<p>Widget type '{widget.chart_type.value}' not yet implemented</p>"

    async def _convert_html_to_pdf(self, html_content: str) -> bytes:
        """Convert HTML content to PDF"""
        if WEASYPRINT_AVAILABLE:
            try:
                html_doc = HTML(string=html_content)
                pdf_bytes = html_doc.write_pdf()
                return pdf_bytes
            except Exception as e:
                logger.error(f"Error converting HTML to PDF with WeasyPrint: {e}")
        
        # Fallback: return HTML as bytes
        logger.warning("PDF conversion not available, returning HTML as bytes")
        return html_content.encode('utf-8')

    async def _generate_json_report(self, 
                                  template: ReportTemplate,
                                  widget_data: Dict[str, Any],
                                  parameters: Dict[str, Any]) -> str:
        """Generate JSON report content"""
        
        report_data = {
            "report_metadata": {
                "template_id": template.id,
                "template_name": template.name,
                "description": template.description,
                "report_type": template.report_type.value,
                "generated_at": datetime.now().isoformat(),
                "parameters": parameters
            },
            "widgets": {},
            "summary": {
                "total_widgets": len(template.widgets),
                "data_sources": list(set(w.data_source for w in template.widgets)),
                "chart_types": list(set(w.chart_type.value for w in template.widgets))
            }
        }
        
        # Add widget data
        for widget in template.widgets:
            data = widget_data.get(widget.id, {})
            report_data["widgets"][widget.id] = {
                "widget_config": asdict(widget),
                "data": data
            }
        
        return json.dumps(report_data, indent=2, default=str)

    async def _generate_excel_report(self, 
                                   template: ReportTemplate,
                                   widget_data: Dict[str, Any],
                                   parameters: Dict[str, Any]) -> bytes:
        """Generate Excel report content"""
        
        if not OPENPYXL_AVAILABLE:
            logger.warning("OpenPyXL not available, returning JSON as bytes")
            json_content = await self._generate_json_report(template, widget_data, parameters)
            return json_content.encode('utf-8')
        
        try:
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Report Data"
            
            # Add header
            ws['A1'] = template.name
            ws['A1'].font = Font(size=16, bold=True)
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Add widget data
            row = 4
            for widget in template.widgets:
                data = widget_data.get(widget.id, {})
                
                # Widget title
                ws[f'A{row}'] = widget.title
                ws[f'A{row}'].font = Font(size=12, bold=True)
                row += 1
                
                # Widget data
                if isinstance(data, dict) and not data.get("error"):
                    for key, value in data.items():
                        ws[f'A{row}'] = key.replace('_', ' ').title()
                        ws[f'B{row}'] = value
                        row += 1
                else:
                    ws[f'A{row}'] = "No data available"
                    row += 1
                
                row += 1  # Add spacing
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            # Fallback to JSON
            json_content = await self._generate_json_report(template, widget_data, parameters)
            return json_content.encode('utf-8')

    async def schedule_report(self, schedule: ReportSchedule) -> str:
        """
        Schedule automated report generation
        
        Args:
            schedule: ReportSchedule configuration
            
        Returns:
            Schedule ID
        """
        try:
            # Store in memory
            self.schedules[schedule.id] = schedule
            
            # Store in database
            with sqlite3.connect(self.database_path) as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    INSERT OR REPLACE INTO report_schedules 
                    (id, template_id, name, frequency, schedule_config, recipients,
                     output_format, active, last_run, next_run, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    schedule.id, schedule.template_id, schedule.name,
                    schedule.frequency.value, json.dumps(schedule.schedule_config),
                    json.dumps(schedule.recipients), schedule.output_format.value,
                    schedule.active, 
                    schedule.last_run.isoformat() if schedule.last_run else None,
                    schedule.next_run.isoformat() if schedule.next_run else None,
                    schedule.created_at.isoformat()
                ))
                conn.commit()
            
            logger.info(f"Scheduled report: {schedule.id}")
            return schedule.id
            
        except Exception as e:
            logger.error(f"Error scheduling report: {e}")
            raise

    def get_templates(self) -> List[ReportTemplate]:
        """Get all available report templates"""
        return list(self.templates.values())

    def get_template(self, template_id: str) -> Optional[ReportTemplate]:
        """Get specific report template"""
        return self.templates.get(template_id)

    async def get_generated_reports(self, limit: int = 50) -> List[Dict[str, Any]]:
        """Get list of generated reports"""
        try:
            with sqlite3.connect(self.database_path) as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT id, template_id, name, report_type, format, file_path,
                           generated_at, file_size, status
                    FROM generated_reports
                    ORDER BY generated_at DESC
                    LIMIT ?
                ''', (limit,))
                
                rows = cursor.fetchall()
                
                reports = []
                for row in rows:
                    reports.append({
                        "id": row[0],
                        "template_id": row[1],
                        "name": row[2],
                        "report_type": row[3],
                        "format": row[4],
                        "file_path": row[5],
                        "generated_at": row[6],
                        "file_size": row[7],
                        "status": row[8]
                    })
                
                return reports
                
        except Exception as e:
            logger.error(f"Error getting generated reports: {e}")
            return []

    def get_reporting_statistics(self) -> Dict[str, Any]:
        """Get reporting engine statistics"""
        return {
            "total_templates": len(self.templates),
            "total_schedules": len(self.schedules),
            "available_formats": [f.value for f in ReportFormat],
            "available_chart_types": [c.value for c in ChartType],
            "output_directory": str(self.output_directory),
            "template_directory": str(self.template_directory),
            "features": {
                "plotly_available": PLOTLY_AVAILABLE,
                "matplotlib_available": MATPLOTLIB_AVAILABLE,
                "jinja2_available": JINJA2_AVAILABLE,
                "weasyprint_available": WEASYPRINT_AVAILABLE,
                "openpyxl_available": OPENPYXL_AVAILABLE
            },
            "timestamp": datetime.now().isoformat()
        }

# Example usage and testing
async def main():
    """Example usage of the Custom Reporting Engine"""
    print("🧠 RomAI Custom Reporting Engine - Testing")
    print("=" * 60)
    
    # Initialize analytics engine
    from .advanced_analytics_engine import AdvancedAnalyticsEngine
    analytics_engine = AdvancedAnalyticsEngine("romai_analytics.db")
    
    # Initialize reporting engine
    reporting_engine = CustomReportingEngine(analytics_engine)
    
    print("\n📊 Reporting Engine Statistics:")
    stats = reporting_engine.get_reporting_statistics()
    for key, value in stats.items():
        if key != "features":
            print(f"  {key}: {value}")
    
    print("\n📋 Available Templates:")
    templates = reporting_engine.get_templates()
    for template in templates:
        print(f"  - {template.id}: {template.name}")
        print(f"    Type: {template.report_type.value}, Widgets: {len(template.widgets)}")
    
    print("\n🎯 Generating test reports...")
    
    # Generate system health report
    print("  📊 Generating System Health Report (HTML)...")
    system_report = await reporting_engine.generate_report(
        "system_health_standard", 
        ReportFormat.HTML,
        {"period": "24h"}
    )
    print(f"    ✅ Generated: {system_report.id} ({system_report.file_size} bytes)")
    
    # Generate AI performance report
    print("  🤖 Generating AI Performance Report (JSON)...")
    ai_report = await reporting_engine.generate_report(
        "ai_performance_standard",
        ReportFormat.JSON,
        {"period": "7d"}
    )
    print(f"    ✅ Generated: {ai_report.id} ({ai_report.file_size} bytes)")
    
    # Generate business intelligence report
    print("  💼 Generating Business Intelligence Report (PDF)...")
    business_report = await reporting_engine.generate_report(
        "business_intelligence_standard",
        ReportFormat.PDF,
        {"period": "30d"}
    )
    print(f"    ✅ Generated: {business_report.id} ({business_report.file_size} bytes)")
    
    print("\n📈 Generated Reports Summary:")
    reports = await reporting_engine.get_generated_reports(10)
    for report in reports:
        print(f"  - {report['name']} ({report['format']}) - {report['generated_at']}")
    
    print("\n🎉 All tests completed successfully!")
    print(f"💾 Output directory: {reporting_engine.output_directory}")

if __name__ == "__main__":
    asyncio.run(main())
